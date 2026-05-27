from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk
from openpyxl import Workbook, load_workbook


FUNDO = "#101010"
CARD = "#1b1b1b"
BOTAO = "#242424"
HOVER = "#323232"
AZUL = "#1f6aa5"
AZUL_HOVER = "#144870"


class ExcelFrame(ctk.CTkFrame):

    def __init__(self, master):

        super().__init__(
            master,
            fg_color=FUNDO
        )

        self.caminho_arquivo = ""
        self.workbook = None
        self.planilha = None
        self.linhas_widgets = {}
        self.alteracoes_pendentes = False

        self.titulo = ctk.CTkLabel(
            self,
            text="Atualização de Preços",
            font=("Segoe UI", 28, "bold")
        )

        self.titulo.pack(pady=(18, 8))

        self.frame_arquivo = ctk.CTkFrame(
            self,
            fg_color="transparent"
        )

        self.frame_arquivo.pack(pady=(0, 8))

        self.btn_novo = ctk.CTkButton(
            self.frame_arquivo,
            text="Nova Planilha",
            command=self.criar_planilha,
            fg_color=BOTAO,
            hover_color=HOVER,
            height=40,
            width=160
        )

        self.btn_novo.grid(
            row=0,
            column=0,
            padx=8
        )

        self.btn_arquivo = ctk.CTkButton(
            self.frame_arquivo,
            text="Selecionar Planilha",
            command=self.selecionar_arquivo,
            fg_color=BOTAO,
            hover_color=HOVER,
            height=40,
            width=180
        )

        self.btn_arquivo.grid(
            row=0,
            column=1,
            padx=8
        )

        self.label_arquivo = ctk.CTkLabel(
            self,
            text="Nenhum arquivo selecionado",
            text_color="gray"
        )

        self.label_arquivo.pack(pady=(0, 10))

        self.frame_inputs = ctk.CTkFrame(
            self,
            fg_color=CARD,
            corner_radius=15
        )

        self.frame_inputs.pack(
            pady=6,
            padx=30
        )

        self.entry_produto = ctk.CTkEntry(
            self.frame_inputs,
            placeholder_text="Nome do produto",
            width=220,
            height=40
        )

        self.entry_produto.grid(
            row=0,
            column=0,
            padx=10,
            pady=(15, 8)
        )

        self.entry_preco = ctk.CTkEntry(
            self.frame_inputs,
            placeholder_text="Preço",
            width=150,
            height=40
        )

        self.entry_preco.grid(
            row=0,
            column=1,
            padx=10,
            pady=(15, 8)
        )

        self.btn_adicionar = ctk.CTkButton(
            self.frame_inputs,
            text="Adicionar",
            command=self.adicionar_produto,
            fg_color=AZUL,
            hover_color=AZUL_HOVER,
            width=140,
            height=40
        )

        self.btn_adicionar.grid(
            row=0,
            column=2,
            padx=10,
            pady=(15, 8)
        )

        self.entry_percentual = ctk.CTkEntry(
            self.frame_inputs,
            placeholder_text="% alteração",
            width=150,
            height=40
        )

        self.entry_percentual.grid(
            row=1,
            column=0,
            padx=10,
            pady=(0, 15)
        )

        self.btn_aumentar = ctk.CTkButton(
            self.frame_inputs,
            text="Aumentar %",
            command=self.aumentar_percentual,
            fg_color=BOTAO,
            hover_color=HOVER,
            width=140,
            height=40
        )

        self.btn_aumentar.grid(
            row=1,
            column=1,
            padx=10,
            pady=(0, 15)
        )

        self.btn_diminuir = ctk.CTkButton(
            self.frame_inputs,
            text="Diminuir %",
            command=self.diminuir_percentual,
            fg_color=BOTAO,
            hover_color=HOVER,
            width=140,
            height=40
        )

        self.btn_diminuir.grid(
            row=1,
            column=2,
            padx=10,
            pady=(0, 15)
        )

        self.frame_acoes = ctk.CTkFrame(
            self,
            fg_color="transparent"
        )

        self.frame_acoes.pack(pady=(8, 10))

        self.btn_aplicar = ctk.CTkButton(
            self.frame_acoes,
            text="Editar Produto",
            command=self.editar_produto_por_nome,
            fg_color=BOTAO,
            hover_color=HOVER,
            width=150,
            height=40
        )

        self.btn_aplicar.grid(
            row=0,
            column=0,
            padx=7
        )

        self.btn_salvar_atual = ctk.CTkButton(
            self.frame_acoes,
            text="Salvar Alterações",
            command=self.salvar_alteracoes,
            fg_color=BOTAO,
            hover_color=HOVER,
            width=160,
            height=40
        )

        self.btn_salvar_atual.grid(
            row=0,
            column=1,
            padx=7
        )

        self.btn_salvar = ctk.CTkButton(
            self.frame_acoes,
            text="Salvar Como",
            command=self.salvar_planilha,
            fg_color=BOTAO,
            hover_color=HOVER,
            width=140,
            height=40
        )

        self.btn_salvar.grid(
            row=0,
            column=2,
            padx=7
        )

        self.label_status = ctk.CTkLabel(
            self,
            text="",
            text_color="gray",
            font=("Segoe UI", 13)
        )

        self.label_status.pack(pady=(0, 8))

        self.tabela = ctk.CTkScrollableFrame(
            self,
            width=660,
            height=280,
            fg_color=CARD,
            corner_radius=15
        )

        self.tabela.pack(
            padx=20,
            pady=(0, 20)
        )

        self.entry_produto.bind(
            "<Return>",
            lambda evento: self.adicionar_produto()
        )
        self.entry_preco.bind(
            "<Return>",
            lambda evento: self.adicionar_produto()
        )

        self.carregar_planilha_padrao()

    def carregar_planilha_padrao(self):
        caminho_padrao = (
            Path(__file__).resolve().parents[1]
            / "Arquivos"
            / "precos.xlsx"
        )

        if caminho_padrao.exists():
            self.carregar_arquivo(caminho_padrao)

        else:
            self.atualizar_tabela()

    def selecionar_arquivo(self):
        arquivo = filedialog.askopenfilename(
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )

        if not arquivo:
            return

        self.carregar_arquivo(arquivo)

    def carregar_arquivo(self, arquivo):
        try:
            self.workbook = load_workbook(arquivo)
            self.planilha = self.workbook.active
            self.caminho_arquivo = str(arquivo)
            self.alteracoes_pendentes = False

            self.garantir_cabecalho()
            self.label_arquivo.configure(
                text=str(arquivo),
                text_color="gray"
            )
            self.exibir_mensagem("Planilha carregada")
            self.atualizar_tabela()

        except Exception as erro:
            self.workbook = None
            self.planilha = None
            self.caminho_arquivo = ""
            self.alteracoes_pendentes = False
            self.label_arquivo.configure(
                text="Erro ao carregar a planilha",
                text_color="red"
            )
            self.exibir_mensagem(
                f"Erro ao carregar a planilha: {erro}",
                "red"
            )
            self.atualizar_tabela()

    def criar_planilha(self):
        self.workbook = Workbook()
        self.planilha = self.workbook.active
        self.planilha.title = "Precos"
        self.caminho_arquivo = ""
        self.alteracoes_pendentes = False

        self.garantir_cabecalho()
        self.limpar_campos()
        self.label_arquivo.configure(
            text="Nova planilha ainda não salva",
            text_color="gray"
        )
        self.exibir_mensagem(
            "Nova planilha criada. Adicione itens e use Salvar Como."
        )
        self.atualizar_tabela()

    def garantir_cabecalho(self):
        if not self.planilha:
            return

        if self.planilha["A1"].value is None:
            self.planilha["A1"] = "Produto"

        if self.planilha["B1"].value is None:
            self.planilha["B1"] = "Preco"

    def atualizar_tabela(self):
        for widget in self.tabela.winfo_children():
            widget.destroy()

        self.linhas_widgets = {}

        self.criar_cabecalho_tabela()

        if not self.planilha:
            self.criar_linha_vazia(
                "Nenhuma planilha carregada"
            )
            return

        linha_tela = 1

        for linha in range(2, self.planilha.max_row + 1):
            produto = self.planilha[f"A{linha}"].value
            preco = self.planilha[f"B{linha}"].value

            if produto is None and preco is None:
                continue

            self.criar_linha_tabela(
                linha_tela,
                linha,
                produto,
                preco
            )
            linha_tela += 1

        if linha_tela == 1:
            self.criar_linha_vazia(
                "Nenhum produto cadastrado"
            )

    def criar_cabecalho_tabela(self):
        colunas = [
            ("#", 45),
            ("Produto", 250),
            ("Preco", 115),
            ("Editar", 82),
            ("Remover", 92),
        ]

        for coluna, (texto, largura) in enumerate(colunas):
            label = ctk.CTkLabel(
                self.tabela,
                text=texto,
                width=largura,
                font=("Segoe UI", 13, "bold"),
                text_color="gray"
            )
            label.grid(
                row=0,
                column=coluna,
                padx=4,
                pady=(8, 6),
                sticky="ew"
            )

    def criar_linha_tabela(self, linha_tela, linha_planilha, produto, preco):
        numero = ctk.CTkLabel(
            self.tabela,
            text=str(linha_planilha - 1),
            width=45
        )
        numero.grid(
            row=linha_tela,
            column=0,
            padx=4,
            pady=5
        )

        entry_produto = ctk.CTkEntry(
            self.tabela,
            width=250,
            height=38
        )
        entry_produto.insert(
            0,
            "" if produto is None else str(produto)
        )
        entry_produto.grid(
            row=linha_tela,
            column=1,
            padx=4,
            pady=5,
            sticky="ew"
        )

        entry_preco = ctk.CTkEntry(
            self.tabela,
            width=115,
            height=38
        )
        entry_preco.insert(
            0,
            self.formatar_numero_edicao(preco)
        )
        entry_preco.grid(
            row=linha_tela,
            column=2,
            padx=4,
            pady=5
        )

        btn_salvar = ctk.CTkButton(
            self.tabela,
            text="Editar",
            command=lambda linha=linha_planilha: self.salvar_linha(linha),
            fg_color=BOTAO,
            hover_color=HOVER,
            width=82,
            height=38
        )
        btn_salvar.grid(
            row=linha_tela,
            column=3,
            padx=4,
            pady=5
        )

        btn_remover = ctk.CTkButton(
            self.tabela,
            text="Remover",
            command=lambda linha=linha_planilha: self.remover_linha(linha),
            fg_color=BOTAO,
            hover_color=HOVER,
            width=92,
            height=38
        )
        btn_remover.grid(
            row=linha_tela,
            column=4,
            padx=4,
            pady=5
        )

        entry_produto.bind(
            "<Return>",
            lambda evento, linha=linha_planilha: self.salvar_linha(linha)
        )
        entry_preco.bind(
            "<Return>",
            lambda evento, linha=linha_planilha: self.salvar_linha(linha)
        )

        self.linhas_widgets[linha_planilha] = (
            entry_produto,
            entry_preco
        )

    def criar_linha_vazia(self, texto):
        label = ctk.CTkLabel(
            self.tabela,
            text=texto,
            text_color="gray",
            font=("Segoe UI", 14)
        )
        label.grid(
            row=1,
            column=0,
            columnspan=5,
            padx=12,
            pady=30
        )

    def normalizar_produto(self, produto):
        return str(produto or "").strip().casefold()

    def encontrar_linha_produto(self, produto):
        if not self.planilha:
            return None

        produto_normalizado = self.normalizar_produto(produto)

        if not produto_normalizado:
            return None

        for linha in range(2, self.planilha.max_row + 1):
            produto_planilha = self.planilha[f"A{linha}"].value

            if self.normalizar_produto(produto_planilha) == produto_normalizado:
                return linha

        return None

    def editar_produto_por_nome(self):
        if not self.planilha:
            self.exibir_mensagem("Crie ou selecione uma planilha primeiro", "red")
            return

        produto = self.entry_produto.get().strip()

        if not produto:
            self.exibir_mensagem("Digite o nome do produto para editar", "red")
            self.entry_produto.focus_set()
            return

        try:
            preco = self.converter_numero(
                self.entry_preco.get()
            )

        except ValueError:
            self.exibir_mensagem("Digite um preco valido para editar", "red")
            self.entry_preco.focus_set()
            return

        linha = self.encontrar_linha_produto(produto)

        if not linha:
            self.exibir_mensagem(
                "Produto nao encontrado. Use Adicionar para cadastrar.",
                "red"
            )
            return

        self.planilha[f"A{linha}"] = produto
        self.planilha[f"B{linha}"] = preco
        self.alteracoes_pendentes = True
        self.limpar_campos()
        self.exibir_mensagem(
            f"{produto} editado. Use Salvar Alteracoes para gravar."
        )
        self.atualizar_tabela()

    def adicionar_produto(self):
        if not self.planilha:
            self.exibir_mensagem("Crie ou selecione uma planilha primeiro", "red")
            return

        produto = self.entry_produto.get().strip()

        if not produto:
            self.exibir_mensagem("Digite o nome do produto", "red")
            return

        try:
            preco = self.converter_numero(
                self.entry_preco.get()
            )

        except ValueError:
            self.exibir_mensagem("Digite um preço válido", "red")
            return

        if self.encontrar_linha_produto(produto):
            self.exibir_mensagem(
                "Produto ja existe. Use Editar Produto para alterar.",
                "red"
            )
            return

        nova_linha = self.proxima_linha()

        self.planilha[f"A{nova_linha}"] = produto
        self.planilha[f"B{nova_linha}"] = preco
        self.alteracoes_pendentes = True
        self.limpar_campos()
        self.exibir_mensagem("Produto adicionado")
        self.atualizar_tabela()

    def salvar_linha(
        self,
        linha,
        atualizar_tabela=True,
        mostrar_mensagem=True
    ):
        if not self.planilha or linha not in self.linhas_widgets:
            return False

        entry_produto, entry_preco = self.linhas_widgets[linha]
        produto = entry_produto.get().strip()

        if not produto:
            self.exibir_mensagem(
                f"Linha {linha - 1}: informe o produto",
                "red"
            )
            entry_produto.focus_set()
            return False

        try:
            preco = self.converter_numero(
                entry_preco.get()
            )

        except ValueError:
            self.exibir_mensagem(
                f"Linha {linha - 1}: preço inválido",
                "red"
            )
            entry_preco.focus_set()
            return False

        self.planilha[f"A{linha}"] = produto
        self.planilha[f"B{linha}"] = preco
        self.alteracoes_pendentes = True

        if mostrar_mensagem:
            self.exibir_mensagem(
                f"Linha {linha - 1} editada. Use Salvar Alteracoes para gravar."
            )

        if atualizar_tabela:
            self.atualizar_tabela()

        return True

    def aplicar_edicoes(self, mostrar_mensagem=True):
        if not self.planilha:
            self.exibir_mensagem("Nenhuma planilha carregada", "red")
            return False

        for linha in sorted(self.linhas_widgets):
            if not self.salvar_linha(
                linha,
                atualizar_tabela=False,
                mostrar_mensagem=False
            ):
                return False

        if mostrar_mensagem:
            self.exibir_mensagem("Todas as edições foram aplicadas")

        self.atualizar_tabela()
        return True

    def remover_linha(self, linha):
        if not self.planilha:
            return

        produto = self.planilha[f"A{linha}"].value or f"Linha {linha - 1}"
        self.planilha.delete_rows(linha, 1)
        self.alteracoes_pendentes = True
        self.exibir_mensagem(f"{produto} removido")
        self.atualizar_tabela()

    def aumentar_percentual(self):
        self.alterar_percentual(aumentar=True)

    def diminuir_percentual(self):
        self.alterar_percentual(aumentar=False)

    def alterar_percentual(self, aumentar):
        if not self.planilha:
            self.exibir_mensagem("Crie ou selecione uma planilha primeiro", "red")
            return

        if not self.aplicar_edicoes(mostrar_mensagem=False):
            return

        try:
            percentual = self.converter_numero(
                self.entry_percentual.get()
            )

        except ValueError:
            self.exibir_mensagem("Digite um percentual válido", "red")
            return

        if percentual < 0:
            self.exibir_mensagem("Digite um percentual positivo", "red")
            return

        if not aumentar and percentual > 100:
            self.exibir_mensagem("A redução não pode passar de 100%", "red")
            return

        fator = 1 + percentual / 100 if aumentar else 1 - percentual / 100
        alterados = 0
        ignorados = 0

        for linha in range(2, self.planilha.max_row + 1):
            produto = self.planilha[f"A{linha}"].value
            preco = self.planilha[f"B{linha}"].value

            if produto is None and preco is None:
                continue

            try:
                preco_antigo = self.converter_numero(preco)

            except ValueError:
                ignorados += 1
                continue

            self.planilha[f"B{linha}"] = round(preco_antigo * fator, 2)
            alterados += 1

        if alterados == 0:
            self.exibir_mensagem("Nenhum preço válido encontrado", "red")
            return

        acao = "aumentados" if aumentar else "reduzidos"
        mensagem = f"{alterados} preços {acao}"

        if ignorados:
            mensagem += f"; {ignorados} linhas ignoradas"

        self.exibir_mensagem(mensagem)
        self.alteracoes_pendentes = True
        self.atualizar_tabela()

    def salvar_alteracoes(self):
        if not self.workbook:
            self.exibir_mensagem("Nenhuma planilha carregada", "red")
            return

        if not self.aplicar_edicoes(mostrar_mensagem=False):
            return

        if not self.caminho_arquivo:
            self.salvar_planilha(aplicar=False)
            return

        try:
            self.workbook.save(self.caminho_arquivo)
            self.alteracoes_pendentes = False
            self.exibir_mensagem(
                f"Alteracoes salvas em: {self.caminho_arquivo}"
            )

        except Exception as erro:
            self.exibir_mensagem(
                f"Erro ao salvar a planilha: {erro}",
                "red"
            )

    def salvar_planilha(self, aplicar=True):
        if not self.workbook:
            self.exibir_mensagem("Nenhuma planilha carregada", "red")
            return

        if aplicar and not self.aplicar_edicoes(mostrar_mensagem=False):
            return

        nome_inicial = "precos_atualizados.xlsx"

        if self.caminho_arquivo:
            nome_inicial = Path(self.caminho_arquivo).name

        caminho_salvar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
            initialfile=nome_inicial
        )

        if not caminho_salvar:
            return

        try:
            self.workbook.save(caminho_salvar)
            self.caminho_arquivo = caminho_salvar
            self.alteracoes_pendentes = False
            self.label_arquivo.configure(
                text=caminho_salvar,
                text_color="gray"
            )
            self.exibir_mensagem(
                f"Planilha salva em: {caminho_salvar}"
            )

        except Exception as erro:
            self.exibir_mensagem(
                f"Erro ao salvar a planilha: {erro}",
                "red"
            )

    def proxima_linha(self):
        for linha in range(2, self.planilha.max_row + 1):
            produto = self.planilha[f"A{linha}"].value
            preco = self.planilha[f"B{linha}"].value

            if produto is None and preco is None:
                return linha

        return self.planilha.max_row + 1

    def limpar_campos(self):
        self.entry_produto.delete(
            0,
            "end"
        )
        self.entry_preco.delete(
            0,
            "end"
        )

    def exibir_mensagem(self, mensagem, cor="gray"):
        self.label_status.configure(
            text=mensagem,
            text_color=cor
        )

    def converter_numero(self, valor):
        if valor is None:
            raise ValueError("valor vazio")

        texto = str(valor).strip()

        if not texto:
            raise ValueError("valor vazio")

        texto = (
            texto.replace("R$", "")
            .replace(" ", "")
        )

        if "," in texto and "." in texto:
            texto = texto.replace(".", "")
            texto = texto.replace(",", ".")

        else:
            texto = texto.replace(",", ".")

        return float(texto)

    def formatar_numero_edicao(self, valor):
        if valor is None:
            return ""

        try:
            numero = self.converter_numero(valor)
            return f"{numero:.2f}".replace(".", ",")

        except ValueError:
            return str(valor)

    def formatar_preco(self, valor):
        try:
            numero = self.converter_numero(valor)
            return f"R$ {numero:.2f}".replace(".", ",")

        except ValueError:
            return "preço inválido"
